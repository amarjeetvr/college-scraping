import os
import sys
import json
import time
import asyncio
import random
import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright

# Settings
excel_path = r"E:\amar\development\other work\college scrap\All colleges data.xlsx"
logos_dir = r"E:\amar\development\other work\college scrap\logos"
SAVE_INTERVAL = 50   # Save every 50 processed rows
CONCURRENCY = 8      # Number of concurrent browser tabs

# Metrics
processed_count = 0
scraped_count = 0
failed_count = 0
save_counter = 0

# Create logos directory if not exists
if not os.path.exists(logos_dir):
    os.makedirs(logos_dir)

def download_and_resize_logo(college_id, logo_url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }
        resp = requests.get(logo_url, headers=headers, timeout=5)
        if resp.status_code == 200:
            ext = ".png"
            if "jpeg" in resp.headers.get("Content-Type", "").lower() or "jpg" in logo_url.lower():
                ext = ".jpg"
            elif "gif" in resp.headers.get("Content-Type", "").lower():
                ext = ".gif"
                
            img_path = os.path.join(logos_dir, f"{int(college_id)}{ext}")
            with open(img_path, "wb") as f:
                f.write(resp.content)
                
            # Resize image to fit inside 50x50 box using Pillow
            with PILImage.open(img_path) as pil_img:
                pil_img = pil_img.convert("RGBA")
                pil_img.thumbnail((50, 50))
                
                # Save as transparent PNG
                final_path = os.path.join(logos_dir, f"{int(college_id)}_resized.png")
                pil_img.save(final_path, "PNG")
            return final_path
    except Exception as e:
        print(f"[ERROR] Downloading logo for ID {college_id}: {e}")
    return None

def extract_college_data(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    script = soup.find("script", id="__NEXT_DATA__")
    if not script:
        return None, None, None
        
    try:
        data = json.loads(script.string)
        page_props_data = data["props"]["initialProps"]["pageProps"]["data"]
    except Exception:
        return None, None, None
        
    # 1. Logo
    logo = "Not Specified"
    try:
        logo_path = page_props_data.get("basic_info", {}).get("logo")
        if logo_path:
            logo = logo_path if logo_path.startswith("http") else f"https://img.collegedunia.com/{logo_path}"
    except Exception:
        pass
        
    # 2. About Description
    about = "Not Specified"
    try:
        desc_html = page_props_data.get("article", {}).get("description", "")
        if desc_html:
            desc_soup = BeautifulSoup(desc_html, "html.parser")
            about_paragraphs = []
            
            for elem in desc_soup.find_all(recursive=False):
                if elem.name in ["h2", "h3", "h4", "table"]:
                    break
                if elem.get_text().strip().lower() in ["table of content", "table of contents"]:
                    break
                if elem.name == "div" and ("table" in str(elem.get('class', '')).lower() or "toc" in str(elem.get('class', '')).lower()):
                    break
                text = elem.get_text().strip()
                if text:
                    about_paragraphs.append(text)
                    
            if not about_paragraphs:
                paragraphs = [p.get_text().strip() for p in desc_soup.find_all("p") if p.get_text().strip()]
                for p in paragraphs:
                    if "table of content" in p.lower() or "highlights" in p.lower() or "ques." in p.lower():
                        break
                    about_paragraphs.append(p)
                    if len(about_paragraphs) >= 2:
                        break
            if about_paragraphs:
                about = "\n\n".join(about_paragraphs[:3])
    except Exception:
        pass
        
    # 3. FAQs
    faqs = "Not Specified"
    try:
        desc_html = page_props_data.get("article", {}).get("description", "")
        if desc_html:
            desc_soup = BeautifulSoup(desc_html, "html.parser")
            faq_soup = desc_soup.find(class_="cdcms_faqs") or desc_soup
            strings = list(faq_soup.stripped_strings)
            faq_list = []
            current_q = None
            current_a_parts = []
            in_answer = False
            
            for s in strings:
                s_lower = s.lower()
                if s_lower.startswith("ques.") or s_lower.startswith("ques:") or s_lower.startswith("ques :"):
                    if current_q:
                        ans_text = " ".join(current_a_parts).strip()
                        faq_list.append((current_q, ans_text if ans_text else "Not Specified"))
                    current_q = s
                    current_a_parts = []
                    in_answer = False
                elif s_lower.startswith("ans.") or s_lower.startswith("ans:") or s_lower.startswith("ans :") or s_lower == "ans" or s_lower.startswith("ans "):
                    in_answer = True
                    ans_part = s
                    for prefix in ["ans.", "ans:", "ans :", "ans"]:
                        if ans_part.lower().startswith(prefix):
                            ans_part = ans_part[len(prefix):].strip()
                            break
                    if ans_part:
                        current_a_parts.append(ans_part)
                elif in_answer:
                    if s_lower.startswith("read more:") or "table of content" in s_lower:
                        break
                    current_a_parts.append(s)
                    
            if current_q:
                ans_text = " ".join(current_a_parts).strip()
                faq_list.append((current_q, ans_text if ans_text else "Not Specified"))
                
            if faq_list:
                formatted_faqs = []
                for q, a in faq_list:
                    q_clean = q
                    if not (q_clean.startswith("Ques.") or q_clean.startswith("Ques:") or q_clean.startswith("Ques :")):
                        q_clean = f"Ques. {q_clean}"
                    a_clean = a
                    if a_clean != "Not Specified":
                        if not (a_clean.startswith("Ans.") or a_clean.startswith("Ans:") or a_clean.startswith("Ans :")):
                            a_clean = f"Ans. {a_clean}"
                    formatted_faqs.append(f"{q_clean}\n{a_clean}")
                faqs = "\n\n".join(formatted_faqs)
    except Exception:
        pass
        
    return logo, about, faqs

async def check_internet_connection():
    try:
        loop = asyncio.get_running_loop()
        response = await loop.run_in_executor(None, lambda: requests.get("https://www.google.com", timeout=5))
        return response.status_code == 200
    except Exception:
        return False

async def wait_for_internet():
    print("\n[WARNING] Internet connection lost! Waiting for connection to restore...")
    while not await check_internet_connection():
        print("[WARNING] Connection still down. Checking again in 15 seconds...")
        await asyncio.sleep(15)
    print("[INFO] Internet connection restored! Resuming scraping...\n")

async def main():
    global processed_count, scraped_count, failed_count, save_counter
    
    dry_run = "--dry-run" in sys.argv
    
    print("Loading Excel workbook...")
    if not os.path.exists(excel_path):
        bak_path = excel_path + ".bak"
        if os.path.exists(bak_path):
            print(f"[RECOVERY] Main Excel file not found. Auto-restoring from backup: {bak_path}")
            import shutil
            shutil.copy2(bak_path, excel_path)
        else:
            print(f"Error: Excel file not found at {excel_path}")
            return
        
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["Colleges"]
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
        
    # Get header indexes
    col_indices = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            col_indices[val] = col_idx
            
    # Ensure new columns exist
    new_cols = ['college_logo', 'about_description', 'faq']
    for col_name in new_cols:
        if col_name not in col_indices:
            new_col_idx = ws.max_column + 1
            ws.cell(row=1, column=new_col_idx, value=col_name)
            col_indices[col_name] = new_col_idx
            print(f"Added column '{col_name}' at index {new_col_idx}")
            
    url_col = col_indices.get('url')
    id_col = col_indices.get('college_id')
    logo_col = col_indices.get('college_logo')
    about_col = col_indices.get('about_description')
    faq_col = col_indices.get('faq')
    
    if not url_col or not id_col:
        print("Error: Could not locate 'url' or 'college_id' columns in sheet.")
        return
        
    # Set logo column width to fit 50px image (around 12 character width)
    logo_col_letter = openpyxl.utils.get_column_letter(logo_col)
    ws.column_dimensions[logo_col_letter].width = 12
    
    # Find all unprocessed rows
    rows_to_process = []
    print("Scanning rows to identify unprocessed colleges...")
    for row_idx in range(2, ws.max_row + 1):
        logo_val = ws.cell(row=row_idx, column=logo_col).value
        about_val = ws.cell(row=row_idx, column=about_col).value
        faq_val = ws.cell(row=row_idx, column=faq_col).value
        
        # A row is considered unprocessed if it has neither logo (text/image) nor description nor faq
        # (Drawings don't represent cell values, so checking logo_val text works).
        is_processed = (logo_val is not None) or (about_val is not None) or (faq_val is not None)
        if not is_processed:
            cell = ws.cell(row=row_idx, column=url_col)
            url = cell.hyperlink.target if cell.hyperlink else cell.value
            col_id = ws.cell(row=row_idx, column=id_col).value
            if url and isinstance(url, str) and url.startswith("http"):
                rows_to_process.append((row_idx, col_id, url))
                
    total_colleges = ws.max_row - 1
    total_unprocessed = len(rows_to_process)
    already_scraped = total_colleges - total_unprocessed
    
    print("\n" + "="*60)
    print("                COLLEGE SCRAPING STATUS")
    print("="*60)
    print(f" Total Colleges in Excel:    {total_colleges}")
    print(f" Already Scraped & Saved:    {already_scraped}")
    print(f" Remaining to Scrape:        {total_unprocessed}")
    print("="*60 + "\n")
    
    if dry_run:
        print("--- Running in DRY-RUN mode. Limiting to 5 rows. ---")
        rows_to_process = rows_to_process[:5]
        total_unprocessed = len(rows_to_process)
        
    if not rows_to_process:
        print("All colleges are already scraped and up-to-date! Nothing to do.")
        return

    # Safe asynchronous atomic workbook saver
    global unsaved_changes
    unsaved_changes = False

    async def save_workbook():
        global unsaved_changes
        if not unsaved_changes:
            return
        temp_path = excel_path + f".tmp_{int(time.time())}.xlsx"
        try:
            print(f"[SAVE] Saving workbook progress to Excel...")
            # Save completely to a separate temp file
            await asyncio.to_thread(wb.save, temp_path)
            
            # Update backup
            bak_path = excel_path + ".bak"
            if os.path.exists(excel_path):
                try:
                    import shutil
                    shutil.copy2(excel_path, bak_path)
                except Exception:
                    pass
            
            # Atomically replace destination file
            while True:
                try:
                    os.replace(temp_path, excel_path)
                    print(f"[SAVE] Progress successfully saved to {excel_path}!")
                    unsaved_changes = False
                    break
                except PermissionError:
                    print(f"\n[WARNING] {excel_path} is currently open in Excel and cannot be written to.")
                    print("Please close the Excel file so the script can save progress. Retrying in 15 seconds...\n")
                    await asyncio.sleep(15)
        except Exception as e:
            print(f"[ERROR] Failed to save workbook: {e}")
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    async def update_row_data(row_idx, logo, about, faq, local_logo_path=None):
        global processed_count, unsaved_changes
        ws.cell(row=row_idx, column=logo_col, value=logo)
        ws.cell(row=row_idx, column=about_col, value=about)
        ws.cell(row=row_idx, column=faq_col, value=faq)
        
        # Embed local logo drawing if available
        if local_logo_path and os.path.exists(local_logo_path):
            img = OpenpyxlImage(local_logo_path)
            cell_coord = f"{logo_col_letter}{row_idx}"
            ws.add_image(img, cell_coord)
            ws.row_dimensions[row_idx].height = 45
            
        processed_count += 1
        unsaved_changes = True

    # Async worker task
    semaphore = asyncio.Semaphore(CONCURRENCY)
    
    async def process_row(context, row_info):
        global scraped_count, failed_count
        row_idx, col_id, url = row_info
        
        async with semaphore:
            page = None
            max_row_retries = 3
            for attempt in range(1, max_row_retries + 1):
                try:
                    # Minimal stagger sleep for maximum throughput
                    await asyncio.sleep(random.uniform(0.05, 0.15))
                    
                    if not page:
                        page = await context.new_page()
                    
                    try:
                        await page.goto(url, wait_until="commit", timeout=20000)
                    except Exception:
                        pass
                    
                    # Wait for WAF challenge to solve and __NEXT_DATA__ script tag to attach to the DOM
                    await page.wait_for_selector("script#__NEXT_DATA__", state="attached", timeout=20000)
                    
                    content = await page.content()
                    logo, about, faq = extract_college_data(content)
                    
                    if logo is None:
                        print(f"[FAIL] Row {row_idx} (ID {col_id}): Failed to parse JSON script block")
                        await update_row_data(row_idx, "Not Specified", "Not Specified", "Not Specified")
                        failed_count += 1
                        break
                        
                    # Download image and resize if logo is scraped
                    local_logo_path = None
                    if logo != "Not Specified":
                        local_logo_path = await asyncio.to_thread(download_and_resize_logo, col_id, logo)
                        
                    print(f"[OK] Row {row_idx} (ID {col_id}): Logo={'Scraped & Embedded' if local_logo_path else 'None'}, About={len(about)} chars, FAQs={faq.count('Ques.') if faq != 'Not Specified' else 0}")
                    await update_row_data(row_idx, logo, about, faq, local_logo_path)
                    scraped_count += 1
                    break  # Success, exit retry loop!
                    
                except Exception as e:
                    err_msg = str(e).lower()
                    is_net_error = "err_internet_disconnected" in err_msg or "err_connection" in err_msg or "dns" in err_msg or "network" in err_msg or "net::" in err_msg
                    
                    if is_net_error and attempt < max_row_retries:
                        print(f"[WARNING] Network disconnect detected for Row {row_idx} on attempt {attempt}/{max_row_retries}. Pausing and retrying...")
                        if not await check_internet_connection():
                            await wait_for_internet()
                        # Close current page tab to start fresh on next attempt
                        if page:
                            await page.close()
                            page = None
                    else:
                        print(f"[ERROR] Exception processing row {row_idx}: {e}")
                        await update_row_data(row_idx, "Not Specified", "Not Specified", "Not Specified")
                        failed_count += 1
                        break
            if page:
                await page.close()

    # Split rows into batches of 50
    BATCH_SIZE = 5 if dry_run else 50
    batches = [rows_to_process[i:i + BATCH_SIZE] for i in range(0, len(rows_to_process), BATCH_SIZE)]
    total_batches = len(batches)
    
    print(f"Starting high-speed batch scraping in {total_batches} batches of {BATCH_SIZE} colleges (concurrency={CONCURRENCY})...\n")
    start_time = time.time()
    
    async with async_playwright() as p:
        # Launch Chromium in headful mode
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        
        # Block images, fonts, styles, media, and third-party trackers for ultra-fast scraping
        await context.route("**/*", lambda route: route.abort() if route.request.resource_type in ["image", "font", "stylesheet", "media"] or any(t in route.request.url for t in ["google", "facebook", "clarity", "criteo", "doubleclick", "taboola", "outbrain", "hotjar"]) else route.continue_())
        
        try:
            for batch_idx, batch in enumerate(batches, start=1):
                print(f"\n------------------------------------------------------------")
                print(f"Starting Batch {batch_idx}/{total_batches} ({len(batch)} colleges)...")
                print(f"------------------------------------------------------------")
                
                tasks = [process_row(context, r) for r in batch]
                await asyncio.gather(*tasks)
                
                # Save progress after every batch of 50
                print(f"\n[BATCH {batch_idx} FINISHED] Saving batch of {len(batch)} colleges to Excel...")
                await save_workbook()
                current_total_done = already_scraped + processed_count
                print(f"[PROGRESS] Overall: {current_total_done}/{total_colleges} colleges done ({current_total_done/total_colleges*100:.1f}%)\n")
                
        except (KeyboardInterrupt, asyncio.CancelledError):
            print("\n[STOP] Scraping interrupted by user. Saving any pending progress to Excel...")
        finally:
            await context.close()
            await browser.close()
            # Final save of remaining progress if any
            if unsaved_changes:
                await save_workbook()
                
    elapsed = time.time() - start_time
    print("\n========================================")
    print("Scraping Completed!")
    print(f"Time elapsed: {elapsed:.2f} seconds ({elapsed/60:.2f} minutes)")
    print(f"Total processed in this session: {processed_count}")
    print(f"Successfully scraped: {scraped_count}")
    print(f"Failed to scrape: {failed_count}")
    print(f"Total colleges completed so far: {already_scraped + processed_count}/{total_colleges}")
    print("========================================")

if __name__ == "__main__":
    # Handle event loop execution correctly
    asyncio.run(main())
