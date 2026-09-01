import os
import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import json
import time

excel_path = r"E:\amar\development\other work\college scrap\All colleges data.xlsx"
logos_dir = r"E:\amar\development\other work\college scrap\logos"

# Create logos directory if not exists
if not os.path.exists(logos_dir):
    os.makedirs(logos_dir)
    print(f"Created directory: {logos_dir}")

def fetch_logo_url_with_playwright(url):
    print(f"Launching headful browser to fetch: {url}")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        try:
            response = page.goto(url, timeout=30000)
            if response and response.status == 200:
                soup = BeautifulSoup(page.content(), "html.parser")
                script = soup.find("script", id="__NEXT_DATA__")
                if script:
                    data = json.loads(script.string)
                    logo_path = data["props"]["initialProps"]["pageProps"]["data"].get("basic_info", {}).get("logo")
                    if logo_path:
                        logo_url = logo_path if logo_path.startswith("http") else f"https://img.collegedunia.com/{logo_path}"
                        return logo_url
        except Exception as e:
            print(f"Error fetching URL: {e}")
        finally:
            browser.close()
    return None

def download_and_resize_logo(college_id, logo_url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }
        resp = requests.get(logo_url, headers=headers, timeout=10)
        if resp.status_code == 200:
            # Detect extension from URL or content-type
            ext = ".png"
            if "jpeg" in resp.headers.get("Content-Type", "").lower() or "jpg" in logo_url.lower():
                ext = ".jpg"
            elif "gif" in resp.headers.get("Content-Type", "").lower():
                ext = ".gif"
                
            img_path = os.path.join(logos_dir, f"{int(college_id)}{ext}")
            with open(img_path, "wb") as f:
                f.write(resp.content)
                
            # Resize image to standard size (e.g. 50x50 pixels)
            with PILImage.open(img_path) as pil_img:
                # Convert to RGBA/RGB to support transparency if resizing png
                pil_img = pil_img.convert("RGBA")
                pil_img.thumbnail((50, 50))
                
                # Save as PNG to preserve transparent background
                final_path = os.path.join(logos_dir, f"{int(college_id)}_resized.png")
                pil_img.save(final_path, "PNG")
                
            print(f"Successfully downloaded and resized logo to: {final_path}")
            return final_path
    except Exception as e:
        print(f"Error downloading/resizing logo: {e}")
    return None

def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Colleges"]
    
    # Get column mapping
    col_indices = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            col_indices[val] = col_idx
            
    url_col = col_indices.get('url')
    id_col = col_indices.get('college_id')
    logo_col = col_indices.get('college_logo')
    
    # Set logo column width to fit 50px image (around 10 character width)
    logo_col_letter = openpyxl.utils.get_column_letter(logo_col)
    ws.column_dimensions[logo_col_letter].width = 12
    
    # Process the first 3 data rows (Rows 2, 3, 4)
    rows_to_process = [2, 3, 4]
    
    for row_idx in rows_to_process:
        college_id = ws.cell(row=row_idx, column=id_col).value
        cell = ws.cell(row=row_idx, column=url_col)
        url = cell.hyperlink.target if cell.hyperlink else cell.value
        
        # Check if we already have the logo URL text written in the cell
        logo_url = ws.cell(row=row_idx, column=logo_col).value
        
        print(f"\nProcessing row {row_idx} (ID {college_id})...")
        if logo_url and isinstance(logo_url, str) and logo_url.startswith("http"):
            print(f"Found existing logo URL: {logo_url}")
        else:
            print("Logo URL not in sheet, fetching with Playwright...")
            logo_url = fetch_logo_url_with_playwright(url)
            if logo_url:
                print(f"Fetched logo URL: {logo_url}")
                ws.cell(row=row_idx, column=logo_col, value=logo_url)
            else:
                print("Failed to fetch logo URL.")
                continue
                
        # Download and resize
        local_img_path = download_and_resize_logo(college_id, logo_url)
        if local_img_path:
            # Embed image in Excel
            img = OpenpyxlImage(local_img_path)
            cell_coord = f"{logo_col_letter}{row_idx}"
            ws.add_image(img, cell_coord)
            
            # Increase row height to fit the image nicely (approx 45 points)
            ws.row_dimensions[row_idx].height = 45
            print(f"Embedded image in cell {cell_coord} and set row height to 45.")
            
    print("\nSaving workbook...")
    temp_path = excel_path + ".tmp"
    while True:
        try:
            wb.save(temp_path)
            if os.path.exists(excel_path):
                os.remove(excel_path)
            os.rename(temp_path, excel_path)
            print("Workbook saved successfully!")
            break
        except PermissionError:
            print(f"\n[WARNING] {excel_path} is currently open in Excel and cannot be written to.")
            print("Please close the Excel file to let the script save progress. Retrying in 10 seconds...\n")
            time.sleep(10)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
    wb.close()
    print("Test run completed successfully!")

if __name__ == "__main__":
    main()
